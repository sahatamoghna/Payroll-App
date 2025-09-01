import os, sys, datetime
import pandas as pd
from openpyxl import load_workbook
from utils import get_reports_folder_path
from db_manager import fetch_tax_slabs

# Load template from RESOURCE_DIR when frozen, else code directory
BASE_DIR      = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(BASE_DIR, 'templates', 'payroll_template.xlsx')

# Number formats
CURRENCY_FMT = '#,##0.00'
INTEGER_FMT  = '0'


def compute_paye_amount(taxable: float, slabs: list) -> float:
    """
    Compute PAYE entirely in Python, per your slab logic.
    """
    tax = 0.0
    for slab in sorted(slabs, key=lambda s: s['lower']):
        low, high, rate = slab['lower'], slab['upper'], slab['rate'] / 100.0
        if taxable <= low:
            break
        cap = taxable if high is None else min(taxable, high)
        portion = cap - low
        if portion > 0:
            tax += portion * rate
    return round(tax, 2)


def generate_payroll_report(input_path: str,
                            month: int,
                            year: int,
                            slabs: list = None) -> str:
    if slabs is None:
        slabs = fetch_tax_slabs()

    # Read input, forcing SSN and Tin as text, numeric blanks → 0
    dtype_map = {
        'Social Security No.': str,
        'Tin No.': str
    }
    df = pd.read_excel(input_path, dtype=dtype_map)

    expected = [
        'Name',
        'STAFF ID',
        'Social Security No.',
        'Tin No.',
        'POSITION',
        'Basic Salary',
        'Allowances',
        'OT',
        'Loan Repayment'
    ]
    missing = set(expected) - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    # Replace NaN in numeric columns with 0
    for col in ['Basic Salary', 'Allowances', 'OT', 'Loan Repayment']:
        df[col].fillna(0, inplace=True)

    # Load template
    wb = load_workbook(TEMPLATE_PATH)
    payroll_ws = wb["Payroll Report"]
    summary_ws = wb["Summary"]

    # Header in Payroll Report
    mon = datetime.date(year, month, 1).strftime('%b').upper()
    payroll_ws['B2'] = f"PAYROLL REGISTER FOR THE MONTH OF {mon} {year}"
    payroll_ws['P1'] = datetime.datetime.now().strftime('%d/%m/%Y')

    # Rates from template (shifted positions)
    ssf_rate = payroll_ws['K4'].value or 0.0
    emp_rate = payroll_ws['Q4'].value or 0.0

    # Populate detail rows (row 6 onwards)
    start_row = 6
    for idx, rec in enumerate(df.itertuples(index=False), start_row):
        (name, sid, ssn, tin, pos,
         basic, allow, ot, loan) = rec

        gross      = basic + allow + ot
        ssf_amt    = round(basic * ssf_rate, 2)
        taxable    = basic + ot - ssf_amt
        paye_amt   = compute_paye_amount(taxable, slabs)
        emp_cont   = round(basic * emp_rate, 2)
        deductions = round(paye_amt + ssf_amt + loan, 2)
        net_pay    = round(gross - deductions, 2)

        # Write columns A–T:
        payroll_ws.cell(row=idx, column=1,  value=idx - start_row + 1).number_format = INTEGER_FMT  # A
        payroll_ws.cell(row=idx, column=2,  value=name)                   # B
        payroll_ws.cell(row=idx, column=3,  value=sid)                    # C
        payroll_ws.cell(row=idx, column=4,  value=ssn)                    # D
        payroll_ws.cell(row=idx, column=5,  value=tin)                    # E
        payroll_ws.cell(row=idx, column=6,  value=pos)                    # F

        payroll_ws.cell(row=idx, column=7,  value=basic).number_format     = CURRENCY_FMT  # G
        payroll_ws.cell(row=idx, column=8,  value=allow).number_format     = CURRENCY_FMT  # H
        payroll_ws.cell(row=idx, column=9,  value=ot).number_format        = CURRENCY_FMT  # I
        payroll_ws.cell(row=idx, column=10, value=gross).number_format     = CURRENCY_FMT  # J
        payroll_ws.cell(row=idx, column=11, value=ssf_amt).number_format   = CURRENCY_FMT  # K
        payroll_ws.cell(row=idx, column=12, value=taxable).number_format   = CURRENCY_FMT  # L
        payroll_ws.cell(row=idx, column=13, value=paye_amt).number_format  = CURRENCY_FMT  # M
        payroll_ws.cell(row=idx, column=14, value=loan).number_format      = CURRENCY_FMT  # N
        payroll_ws.cell(row=idx, column=15, value=deductions).number_format= CURRENCY_FMT  # O
        payroll_ws.cell(row=idx, column=16, value=net_pay).number_format   = CURRENCY_FMT  # P
        payroll_ws.cell(row=idx, column=17, value=emp_cont).number_format  = CURRENCY_FMT  # Q
        payroll_ws.cell(row=idx, column=18, value=ssf_amt + emp_cont).number_format = CURRENCY_FMT  # R

        # New columns: 13.5% of Basic → S ; 5% of Basic → T
        col_S = round(basic * 0.135, 2)
        col_T = round(basic * 0.05, 2)
        payroll_ws.cell(row=idx, column=19, value=col_S).number_format = CURRENCY_FMT  # S
        payroll_ws.cell(row=idx, column=20, value=col_T).number_format = CURRENCY_FMT  # T

    last_row = start_row + len(df) - 1
    sheet = payroll_ws.title

    # Credit values originally at C now go to E
    summary_ws["C3"] = f"=SUM('{sheet}'!G{start_row}:G{last_row})"               # Basic Salary
    summary_ws["C4"] = f"=SUM('{sheet}'!H{start_row}:H{last_row})"               # Allowances
    summary_ws["C5"] = f"=SUM('{sheet}'!I{start_row}:I{last_row})"               # OT
    summary_ws["C6"] = f"=SUM('{sheet}'!Q{start_row}:Q{last_row})"               # Employer Contribution

    # Debit values originally at E now go to C
    summary_ws["E3"] = f"=SUM('{sheet}'!M{start_row}:M{last_row})"               # PAYE
    summary_ws["E4"] = f"=SUM('{sheet}'!N{start_row}:N{last_row})"               # Loan Repayment
    summary_ws["E5"] = f"=(13.5/18.5)*SUM('{sheet}'!R{start_row}:R{last_row})"   # SSNIT
    summary_ws["E6"] = f"=(5/18.5)*SUM('{sheet}'!R{start_row}:R{last_row})"      # TIER 2
    summary_ws["E7"] = f"=SUM('{sheet}'!P{start_row}:P{last_row})"               # Net Salaries

    for cell in ("C3","C4","C5","C6","E3","E4","E5","E6","E7"):
        summary_ws[cell].number_format = CURRENCY_FMT

    # Save workbook
    out_folder = get_reports_folder_path()
    os.makedirs(out_folder, exist_ok=True)
    filename = f"{year}-{month:02d}-payroll.xlsx"
    out_path = os.path.join(out_folder, filename)
    wb.save(out_path)
    return out_path
