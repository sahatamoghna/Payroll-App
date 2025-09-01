# payslip_generator.py

import os
import calendar
import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import load_workbook

from utils import show_error, show_info, get_reports_folder_path, format_timestamp
from db_manager import fetch_reports, insert_payslip

def render(parent):
    for w in parent.winfo_children():
        w.destroy()
    style = ttk.Style(parent)
    style.theme_use('clam')
    style.configure('Content.TFrame', background='#FFFFFF')
    style.configure('Content.TLabel', background='#FFFFFF', font=('Segoe UI', 14))
    style.configure('Action.TButton',
                    font=('Segoe UI', 12, 'bold'),
                    background='#6BBF44', foreground='white',
                    padding=(10, 6))
    style.map('Action.TButton', background=[('active', '#57A738')])
    style.configure('TCombobox',
                    fieldbackground='#FFFFFF', background='#FFFFFF')
    style.map('TCombobox',
              fieldbackground=[('readonly', '#FFFFFF')],
              background=[('readonly', '#FFFFFF')])

    frame = ttk.Frame(parent, style='Content.TFrame', padding=20)
    frame.pack(fill='both', expand=True)
    frame.columnconfigure((0,1,2), weight=1)

    # Row 0: Report selector
    ttk.Label(frame, text="Select Report:", style='Content.TLabel')\
        .grid(row=0, column=0, sticky='e', padx=(0,10))
    reports = fetch_reports()
    rpt_map = {}
    names = []
    for _id, m, y, fn, _ in reports:
        key = f"{calendar.month_abbr[m]} {y}"
        names.append(key)
        rpt_map[key] = (m, y, fn)

    report_cb = ttk.Combobox(frame, values=names, state='readonly', width=20)
    report_cb.grid(row=0, column=1, columnspan=2, sticky='w')

    # Row 1: Date input DD / MM / YYYY
    ttk.Label(frame, text="Date:", style='Content.TLabel')\
        .grid(row=1, column=0, sticky='e', pady=(20,0), padx=(0,10))
    date_frame = ttk.Frame(frame, style='Content.TFrame')
    date_frame.grid(row=1, column=1, columnspan=2, sticky='w', pady=(20,0))

    # Validation functions
    def vday(P):
        # allow empty
        if not P:
            return True
        # only digits, max length 2
        if not P.isdigit() or len(P) > 2:
            return False
        # single‐digit: allow 0–3 (so you can start typing “0”)
        if len(P) == 1:
            return P in '0123'
        # two digits: integer must be 01–31
        val = int(P)
        return 1 <= val <= 31

    def vmon(P):
        if not P:
            return True
        if not P.isdigit() or len(P) > 2:
            return False
        # single‐digit: allow 0–1 (so you can start “0” for months)
        if len(P) == 1:
            return P in '01'
        # two digits: integer must be 01–12
        val = int(P)
        return 1 <= val <= 12
    def vyr(P):
        if not P: return True
        return P.isdigit() and len(P)<=4

    vcmd_d = (parent.register(vday), '%P')
    vcmd_m = (parent.register(vmon), '%P')
    vcmd_y = (parent.register(vyr), '%P')

    day_var = tk.StringVar()
    mon_var = tk.StringVar()
    yr_var  = tk.StringVar()

    # Entries + slashes
    day_e = ttk.Entry(date_frame, textvariable=day_var,
                      width=2, justify='center',
                      validate='key', validatecommand=vcmd_d)
    day_e.grid(row=0, column=0)
    ttk.Label(date_frame, text="/", background='#FFFFFF',
              font=('Segoe UI', 14)).grid(row=0, column=1, padx=2)
    mon_e = ttk.Entry(date_frame, textvariable=mon_var,
                      width=2, justify='center',
                      validate='key', validatecommand=vcmd_m)
    mon_e.grid(row=0, column=2)
    ttk.Label(date_frame, text="/", background='#FFFFFF',
              font=('Segoe UI', 14)).grid(row=0, column=3, padx=2)
    yr_e = ttk.Entry(date_frame, textvariable=yr_var,
                     width=4, justify='center',
                     validate='key', validatecommand=vcmd_y)
    yr_e.grid(row=0, column=4)

    # Placeholder labels underneath
    placeholder_style = {'foreground':'#888888','font':('Segoe UI',9), 'background':'#FFFFFF'}
    ttk.Label(date_frame, text="DD", **placeholder_style).grid(row=1, column=0)
    ttk.Label(date_frame, text="MM", **placeholder_style).grid(row=1, column=2)
    ttk.Label(date_frame, text="YYYY", **placeholder_style).grid(row=1, column=4)

    # Row 2: Generate Payslips 
    def on_generate():
        sel = report_cb.get()
        if not sel:
            show_error("Please select a report."); return

        d, mth, y = day_var.get(), mon_var.get(), yr_var.get()
        if not (len(d)==2 and d.isdigit() and 1<=int(d)<=31):
            show_error("Day must be 01–31."); return
        if not (len(mth)==2 and mth.isdigit() and 1<=int(mth)<=12):
            show_error("Month must be 01–12."); return
        if not (len(y)==4 and y.isdigit()):
            show_error("Year must be 4 digits."); return

        report_date = f"{d}/{mth}/{y}"
        mm, yy, fn = rpt_map[sel]
        rpt_path = os.path.join(get_reports_folder_path(), fn)
        if not os.path.exists(rpt_path):
            show_error(f"Missing file:\n{rpt_path}"); return

        dest = filedialog.askdirectory(title="Save payslips")
        if not dest: return
        out_folder = os.path.join(dest, f"Payslips_{sel.replace(' ','_')}")
        os.makedirs(out_folder, exist_ok=True)

        wb = load_workbook(rpt_path, data_only=True)
        ws = wb.active

        row = 6
        count = 0
        from excel_utils import compute_paye_amount
        from db_manager import fetch_tax_slabs
        slabs = fetch_tax_slabs()

        while ws.cell(row=row, column=1).value is not None:
            ssn    = ws.cell(row=row, column=4).value or ""
            tin    = ws.cell(row=row, column=5).value or ""
            name   = ws.cell(row=row, column=2).value or ""
            staff  = ws.cell(row=row, column=3).value or ""
            basic  = ws.cell(row=row, column=7).value or 0.0
            allow  = ws.cell(row=row, column=8).value or 0.0
            ot     = ws.cell(row=row, column=9).value or 0.0
            ssf    = ws.cell(row=row, column=11).value or 0.0
            taxable= ws.cell(row=row, column=12).value or 0.0
            paye   = ws.cell(row=row, column=13).value or 0.0
            loan   = ws.cell(row=row, column=14).value or 0.0
            empc   = ws.cell(row=row, column=17).value or 0.0
            net    = ws.cell(row=row, column=16).value or 0.0
            p13_5  = ws.cell(row=row, column=19).value or 0.0
            p5     = ws.cell(row=row, column=20).value or 0.0

            fmt = lambda v: f"{v:,.2f}"
            data_map = {
                'Social Security No.': ssn,
                'Tin No.':             tin,
                'Name':                name,
                'STAFF ID':            staff,
                'ReportDate':          report_date,
                'Basic Salary':        fmt(basic),
                'Allowances':          fmt(allow),
                'OT':                  fmt(ot),
                'SSF':                 fmt(ssf),
                'Taxable':             fmt(taxable),
                'PAYE':                fmt(paye),
                'Loan Repayment':      fmt(loan),
                'Emp Cont':            fmt(empc),
                'Net Pay':             fmt(net),
                'Percent13.5':         fmt(p13_5),
                'Percent5':            fmt(p5),
            }

            from pdf_utils import generate_payslip_pdf
            pdf_name = f"{staff}_{sel.replace(' ','_')}.pdf"
            out_pdf = os.path.join(out_folder, pdf_name)
            generate_payslip_pdf(staff, mm, yy, data_map, out_pdf)
            insert_payslip(staff, mm, yy, pdf_name, format_timestamp())

            row += 1
            count += 1

        show_info(f"Generated {count} payslips in:\n{out_folder}")

    gen_btn = ttk.Button(frame, text="Generate Payslips",
                         style='Action.TButton', command=on_generate)
    gen_btn.grid(row=2, column=0, columnspan=3, pady=(30,0))

